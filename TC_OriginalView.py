import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
wrkbk = openpyxl.Workbook()
# to create the sheet with name
sh = wrkbk.create_sheet("Results", 0)
# set the value in row 2 and column 3
sh.cell(row=1,column=1).value = "TestCase_Id"
sh.cell(row=1, column=1).font = Font(bold=True)
sh.cell(row=1, column=2).value = "Description"
sh.cell(row=1, column=2).font = Font(bold=True)
sh.cell(row=1, column=3).value = "Status"
sh.cell(row=1, column=3).font = Font(bold=True)


from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
options = webdriver. ChromeOptions()
options. add_experimental_option("detach", True)
driver = webdriver. Chrome(options=options, service=Service(ChromeDriverManager().install()))
# driver = webdriver.Chrome(executable_path= "C://Users//Sindhura//Desktop//Work//Testing//Automation//chromedriver_win32 (1)//chromedriver.exe")
# driver = webdriver.Chrome(options.add_experimental_option()=Service(ChromeDriverManager().install()))
# TestSteps
# 1. Open browser
# 2. Enter url
# 3. Verify whether page displays Loading original view

sh.cell(row=2, column=1).value = 'TC_ID'
sh.cell(row=2, column=2).value = 'Verifying NPI Search page'

driver.get("http://192.168.2.40:4040/npi-search/c2d731f6-907a-45ac-af28-eb7b355fa03e")

driver.maximize_window()
time.sleep(5)
view = driver.find_element(By.XPATH, "/html/body/div[2]/div/div/div/div/div").text
print(view)
if view == 'Loading Original View':
    print('TC_OriginalView passed')
    sh.cell(row=2, column=3).value ='Passed'
    fill_cell2 = PatternFill(patternType='solid',
                             fgColor='35FC03')
    sh.cell(row=2, column=3).fill = fill_cell2
else:
    print('TC_OriginalView failed')
    sh.cell(row=2, column=3).value = 'Failed'
    fill_cell1 = PatternFill(patternType='solid',
                             fgColor='FC2C03')
    sh.cell(row=2, column=3).fill = fill_cell1
time.sleep(5)
wrkbk.save("C:\\Users\\Sindhura\\Desktop\\Work\\Testing\\Automation\\Data_driven.xlsx")
driver.close()