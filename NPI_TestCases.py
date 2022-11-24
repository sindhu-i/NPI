import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import datetime, get_column_letter
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException

wrkbk = openpyxl.Workbook()
# to create the sheet with name
sh = wrkbk.create_sheet("NPI_UI_AutomationReport", 0)
# set the value in row 2 and column 3
sh.cell(row=1,column=1).value = "TestCase_Id"
sh.cell(row=1, column=1).font = Font(bold=True)
sh.cell(row=1, column=2).value = "Description"
sh.cell(row=1, column=2).font = Font(bold=True)
for column_cells in sh.columns:
    new_column_length = max(len(str(cell.value)) for cell in column_cells)
    new_column_letter = (get_column_letter(column_cells[0].column))
    if new_column_length > 0:
        sh.column_dimensions[new_column_letter].width = new_column_length*1.23
# sh.row_dimensions[1:9].height = 70
# sh.column_dimensions[1:3].width = 30
# now = datetime.datetime.now()
# sh.cell(row=1, column=5).value = 'Run date and time'
# sh.cell(row=1, column=5).font = Font(bold=True)
# date = now.strftime("%m-%d-%Y %H:%M:%S")

sh.cell(row=1, column=3).value = "Status"
sh.cell(row=1, column=3).font = Font(bold=True)
# sh.cell(row=1, column=4).value = "Comments"
# sh.cell(row=1, column=4).font = Font(bold=True)



from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
options = webdriver. ChromeOptions()
options. add_experimental_option("detach", True)
driver = webdriver. Chrome(options=options, service=Service(ChromeDriverManager().install()))
# driver = webdriver.Chrome(executable_path= "C://Users//Sindhura//Desktop//Work//Testing//Automation//chromedriver_win32 (1)//chromedriver.exe")
# driver = webdriver.Chrome(options.add_experimental_option()=Service(ChromeDriverManager().install()))
# **********************************************************
# TC_001_NPI Search_page
# TestSteps
# 1. Open browser
# 2. Enter url
# 3. Verify whether page displays Loading original view

sh.cell(row=2, column=1).value = 'TC_001_NPISearchPage'
sh.cell(row=2, column=2).value = 'Verifying NPI Search page'

driver.get("http://192.168.2.40:4040/npi-search/c2d731f6-907a-45ac-af28-eb7b355fa03e")

driver.maximize_window()
time.sleep(20)
view = driver.find_element(By.XPATH, "//*[@id='root']/div[2]/div[1]/div[2]/div/div[1]/span").text
print(view)
Grid_data = driver.find_element(By.XPATH, "//li[normalize-space()='Total Records: 6']").text
if view == 'Original View' and 'Total Records' in Grid_data:
    print('TC_001_OriginalView passed')
    sh.cell(row=2, column=3).value ='Passed'
    fill_cell2 = PatternFill(patternType='solid',
                             fgColor='35FC03')
    sh.cell(row=2, column=3).fill = fill_cell2
else:
    print('TC_OriginalView failed')
    sh.cell(row=2, column=3).value = 'Failed'
    fill_cell1 = PatternFill(patternType='solid',
                             fgColor='FC2C03')
    sh.cell(row=2, column=3).fill = fill_cell1
time.sleep(5)
print("Executing 2nd TestCase")
# *********************************************************************************************
# TC_002_NPISearch_MandatoryFields
# 1. Open browser
# 2. Enter url
# 3. Click on clear button
# 4. Verify all fields are empty
# 5. Click on search button
# 6. Verify mandatory fields must display error message
sh.cell(row=3, column=1).value = 'TC_002_NPISearch_MandatoryFields'
sh.cell(row=3, column=2).value = 'Verifying Mandatory fields'


# click on clear button
driver.find_element(by=By.XPATH, value = "//button[@class='ant-btn']").click()
time.sleep(10)
# checking if NPI type field is empty
# copying NPI, State, County, Zipcode fields attribute
Npi = driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[1]/div[3]/div[1]/div[1]/div[2]/div[1]").text
time.sleep(2)
State = driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[1]/div[3]/div[1]/div[6]/div[2]/div[1]").text
time.sleep(2)
County = driver.find_element(By.XPATH, "//span[normalize-space()='Select County']").text
time.sleep(2)
Zipcode = driver.find_element(By.XPATH, "//div[8]//div[2]//div[1]").text
time.sleep(5)
print(Npi)
print(State)
print(County)
print(Zipcode)


if Npi == 'Select NPI Type' and State == 'Select State' and County == 'Select County' and Zipcode == 'Select Zipcode':
     print("Fields are empty")
     # if State == 'Select State':
     #     print("State field is empty")
     #     if County == 'Select County':
     #         print("County field is empty")
     #         if Zipcode == 'Select Zipcode':
     #             print("Zipcode is empty")
     #         else:
     #             print("Zipcode is not empty")
     time.sleep(2)
     #
     driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[2]/div[1]/div[2]/button[1]").click()
     time.sleep(5)
     NPI_empty = driver.find_element(By.XPATH, "//p[normalize-space()='Please Select the NPI Type!']").text
     State_empty = driver.find_element(By.XPATH, "//p[normalize-space()='Please Select the State!']").text
     County_empty = driver.find_element(By.XPATH, "//p[normalize-space()='Please Select the County!']").text
     Zipcode_empty = driver.find_element(By.XPATH, "//p[normalize-space()='Please Select the Zipcode!']").text
     Blank_Grid = driver.find_element(By.XPATH, "//span[normalize-space()='Please Search Something...']").text
     time.sleep(5)
     print(NPI_empty)
     print(State_empty)
     print(County_empty)
     print(Zipcode_empty)
     print(Blank_Grid)
     if NPI_empty == 'Please Select the NPI Type!' and State_empty == 'Please Select the State!' and County_empty == 'Please Select the County!'and  Zipcode_empty == 'Please Select the Zipcode!' and Blank_Grid == 'Please Search Something...':
         print('TestCase_step2 passed')
         sh.cell(row=3, column=3).value = 'Passed'
         fill_cell2 = PatternFill(patternType='solid',
                                  fgColor='35FC03')
         sh.cell(row=3, column=3).fill = fill_cell2

     else:
         print('TestCase_step1 failed')
         sh.cell(row=3, column=3).value = 'Failed'
         fill_cell1 = PatternFill(patternType='solid',
                                  fgColor='FC2C03')
         sh.cell(row=3, column=3).fill = fill_cell1
else:
     print("Fields are not empty")
     sh.cell(row=3, column=3).value = 'Failed'
     fill_cell1 = PatternFill(patternType='solid',
                              fgColor='FC2C03')
     sh.cell(row=3, column=3).fill = fill_cell1
     time.sleep(10)
print("Executing 3rd TestCase")
# ***************************************************************************************
# TC_003_TestCase for verifying firstname, lastname, organization name fields functionality when user selects NPI type as 'Hospital'
# TestSteps
# 1. Open the browser
# 2. Enter url
# 3. Verify if NPI type is Hospital, then Firstname and last name fields must be disabled
# 4. Verify if NPI type is PCP or Specialist, then Organization name field must be disabled
sh.cell(row=4, column=1).value = 'TC_003_FirstN_LName_OrgName'
sh.cell(row=4, column=2).value = 'Verifying firstname,lastname,Orgname fields functionality if NPI type is Hospital'

# Click on NPI Type dropdown
driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[1]/div[3]/div[1]/div[1]/div[2]/div[1]").click()
time.sleep(2)
# Select Hospital
driver.find_element(By.XPATH, "//div[@title = 'Hospital']").click()
time.sleep(2)
# # Get NPI type selected value
Npi = driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[1]/div[3]/div[1]/div[1]/div[2]/div[1]").text
time.sleep(2)

print(Npi)
time.sleep(2)
if Npi == 'Hospital':
    # Verifying Firstname and last name fields are disabled
    # F_name = driver.find_element(By.CSS_SELECTOR,"body div[id='root'] div[class='content_content__A0Ufa'] div[class='ant-row content_search-filter-wrapper__ZGGmI'] div[class='content_show__l1TJz'] div[class='ant-col ant-col-24 content_filter-container__PFeCY undefined content_show__l1TJz'] div:nth-child(3) span:nth-child(1)")
    first_name = driver.find_element(By.XPATH, "//input[@placeholder='Enter First Name']")
    last_name = driver.find_element(By.XPATH, "//input[@placeholder='Enter Last Name']")
    # = driver.find_element(By.XPATH, "//body/div/div/div/div/div/div/span[2]").is_enabled()
    # print(L_name)

    fname = first_name.is_enabled()
    print(fname)
    lname = first_name.is_enabled()
    print(lname)
    time.sleep(2)
    # Org_name = driver.find_element(By.XPATH, "//span[@class='ant-input-affix-wrapper']").is_enabled()
    org_name = driver.find_element(By.XPATH, "//input[@placeholder='Enter Organization Name']")
    o = org_name.is_enabled()
    print(o)
    if fname == False and lname == False and o == True:
        print("TestCase passed")
        sh.cell(row=4, column=3).value = 'Passed'
        fill_cell2 = PatternFill(patternType='solid',
                                 fgColor='35FC03')
        sh.cell(row=4, column=3).fill = fill_cell2

    else:
        print("Testcase failed")
        sh.cell(row=4, column=3).value = 'Failed'
        fill_cell1 = PatternFill(patternType='solid',
                                 fgColor='FC2C03')
        sh.cell(row=4, column=3).fill = fill_cell1

else:
    print("Not hospital")
    sh.cell(row=4, column=3).value = 'Failed'
    fill_cell1 = PatternFill(patternType='solid',
                             fgColor='FC2C03')
    sh.cell(row=4, column=3).fill = fill_cell1
time.sleep(3)
# *********************************************************************************
# TC_004
sh.cell(row=5, column=1).value = 'TC_004_FirstN_LName_OrgName'
sh.cell(row=5, column=2).value = 'Verifying firstname,lastname,Orgname fields functionality if NPI Type is PCP'
print("Executing 4th TestCase")
# Clicking on NPI type dropdown
driver.find_element(By.XPATH, "//*[@id='root']/div[2]/div[1]/div[3]/div/div[1]/div[2]").click()
time.sleep(5)
# Selecting PCP
driver.find_element(By.XPATH, "//div[@title = 'PCP']").click()
time.sleep(2)
# Verifying the selected value in NPI type
PCP = driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[1]/div[3]/div[1]/div[1]/div[2]/div[1]").text
print(PCP)
time.sleep(5)
if PCP == 'PCP':
    print("TC_004 1st checkpoint passed")
    first_name = driver.find_element(By.XPATH, "//input[@placeholder='Enter First Name']")
    last_name = driver.find_element(By.XPATH, "//input[@placeholder='Enter Last Name']")
    org_name = driver.find_element(By.XPATH, "//input[@placeholder='Enter Organization Name']")
    print(first_name.is_enabled())
    print(last_name.is_enabled())
    print(org_name.is_enabled())
    time.sleep(2)
    if first_name.is_enabled() == True and last_name.is_enabled() == True and org_name.is_enabled() == False:
        print("TC_004 passed")
        sh.cell(row=5, column=3).value = 'Passed'
        fill_cell2 = PatternFill(patternType='solid',
                                 fgColor='35FC03')
        sh.cell(row=5, column=3).fill = fill_cell2
    else:
        print("TC_004 failed")
        sh.cell(row=5, column=3).value = 'Failed'
        fill_cell1 = PatternFill(patternType='solid',
                                 fgColor='FC2C03')
        sh.cell(row=5, column=3).fill = fill_cell1
else:
    print("TC_004 1st checkpoint failed")
    sh.cell(row=5, column=3).value = 'Failed'
    fill_cell1 = PatternFill(patternType='solid',
                             fgColor='FC2C03')
    sh.cell(row=5, column=3).fill = fill_cell1
time.sleep(3)
# *********************************************************************************

# TC_005
sh.cell(row=6, column=1).value = 'TC_005_FirstN_LName_OrgName'
sh.cell(row=6, column=2).value = 'Verifying firstname,lastname,Orgname fields functionality if NPI type is Specialist'
print("Executing 5th TestCase")
# Clicking on NPI type dropdown
driver.find_element(By.XPATH, "//*[@id='root']/div[2]/div[1]/div[3]/div/div[1]/div[2]").click()
time.sleep(5)
# Selecting Specialist
driver.find_element(By.XPATH, "//div[@title = 'Specialist']").click()
time.sleep(2)
# Verifying the selected value in NPI type
Spec = driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[1]/div[3]/div[1]/div[1]/div[2]/div[1]").text
print(Spec)
time.sleep(5)
if Spec == 'Specialist':
    print("TC_005 1st checkpoint passed")
    first_name = driver.find_element(By.XPATH, "//input[@placeholder='Enter First Name']")
    last_name = driver.find_element(By.XPATH, "//input[@placeholder='Enter Last Name']")
    org_name = driver.find_element(By.XPATH, "//input[@placeholder='Enter Organization Name']")
    print(first_name.is_enabled())
    print(last_name.is_enabled())
    print(org_name.is_enabled())
    time.sleep(2)
    if first_name.is_enabled() == True and last_name.is_enabled() == True and org_name.is_enabled() == False:
        print("TC_005 passed")
        sh.cell(row=6, column=3).value = 'Passed'
        fill_cell2 = PatternFill(patternType='solid',
                                 fgColor='35FC03')
        sh.cell(row=6, column=3).fill = fill_cell2
    else:
        print("TC_005 failed")
        sh.cell(row=6, column=3).value = 'Failed'
        fill_cell1 = PatternFill(patternType='solid',
                                 fgColor='FC2C03')
        sh.cell(row=6, column=3).fill = fill_cell1
else:
    print("TC_005 1st checkpoint failed")
    sh.cell(row=6, column=3).value = 'Failed'
    fill_cell1 = PatternFill(patternType='solid',
                             fgColor='FC2C03')
    sh.cell(row=6, column=3).fill = fill_cell1
time.sleep(3)
# *************************************************************************************************
print("Executing 6th TestCase")
sh.cell(row=7, column=1).value = 'TC_006_AffiliatedToHospital_Dropdown'
sh.cell(row=7, column=2).value = 'Verifying Affiliated to Hospital dropdown if NPI type is Specialist'

# TC_006 Verifying affiliated  to hospital dropdown when NPI type is Specialist
# Checking whether Affiliated to Hospital is present or not
Aff_Hosp_PCP = driver.find_element(By.XPATH, "//span[normalize-space()='Affiliated to Hospital']")
time.sleep(3)
A_Spec = Aff_Hosp_PCP.is_displayed()
print(A_Spec)
if A_Spec == True:
    print("TC_006 passed")
    sh.cell(row=7, column=3).value = 'Passed'
    fill_cell2 = PatternFill(patternType='solid',
                             fgColor='35FC03')
    sh.cell(row=7, column=3).fill = fill_cell2
else:
    print("TC_006 failed")
    sh.cell(row=7, column=3).value = 'Failed'
    fill_cell1 = PatternFill(patternType='solid',
                             fgColor='FC2C03')
    sh.cell(row=7, column=3).fill = fill_cell1
time.sleep(3)
# ******************************************************************************************
print("Executing 7th testcase")
sh.cell(row=8, column=1).value = 'TC_007_AffiliatedToHospital_dropdown'
sh.cell(row=8, column=2).value = 'Verifying Affiliated to Hospital dropdown if NPI type is Hospital'

# CLick NPI type dropdown
driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[1]/div[3]/div[1]/div[1]/div[2]/div[1]").click()
time.sleep(2)
# Select Hospital
driver.find_element(By.XPATH, "//div[@title = 'Hospital']").click()
time.sleep(2)
# # Get NPI type selected value
Npi_hosp = driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[1]/div[3]/div[1]/div[1]/div[2]/div[1]").text
time.sleep(2)
if Npi_hosp == 'Hospital':

    try:
        Aff_Hosp = driver.find_element(By.XPATH, "//span[normalize-space()='Affiliated to Hospital']")
        sh.cell(row=8, column=3).value = 'Failed'
        fill_cell1 = PatternFill(patternType='solid',
                                 fgColor='FC2C03')
        sh.cell(row=8, column=3).fill = fill_cell1

    except NoSuchElementException:  # spelling error making this code not work as expected
        pass

        sh.cell(row=8, column=3).value = 'Passed'
        fill_cell2 = PatternFill(patternType='solid',
                             fgColor='35FC03')
        sh.cell(row=8, column=3).fill = fill_cell2
else:
    print("TC_007 failed")
    sh.cell(row=8, column=3).value = 'Failed'
    fill_cell1 = PatternFill(patternType='solid',
                             fgColor='FC2C03')
    sh.cell(row=8, column=3).fill = fill_cell1
time.sleep(5)
# *****************************************************************************************************
print("Executing 8th TestCase")
sh.cell(row=9, column=1).value = 'TC_008_AffiliatedHospital'
sh.cell(row=9, column=2).value = 'Verifying Affiliated to Hospital dropdown if NPI type is PCP'

# Clicking on NPI type dropdown
driver.find_element(By.XPATH, "//*[@id='root']/div[2]/div[1]/div[3]/div/div[1]/div[2]").click()
time.sleep(5)
# Selecting PCP
driver.find_element(By.XPATH, "//div[@title = 'PCP']").click()
time.sleep(2)
# Verifying the selected value in NPI type
Npi_PCP = driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[1]/div[3]/div[1]/div[1]/div[2]/div[1]").text
print(Npi_PCP)
if Npi_PCP == 'PCP':
    print("Testcase8 1st checkpoint passed")
    Aff_Hosp_PCP = driver.find_element(By.XPATH, "//span[normalize-space()='Affiliated to Hospital']")
    time.sleep(3)
    A_PCP = Aff_Hosp_PCP.is_displayed()
    print(A_PCP)
    if A_PCP == True:
            print("8th Testcase passed")
            sh.cell(row=9, column=3).value = 'Passed'
            fill_cell2 = PatternFill(patternType='solid',
                                     fgColor='35FC03')
            sh.cell(row=9, column=3).fill = fill_cell2
    else:
            print("TC_008 failed")
            sh.cell(row=9, column=3).value = 'Failed'
            fill_cell1 = PatternFill(patternType='solid',
                                     fgColor='FC2C03')
            sh.cell(row=9, column=3).fill = fill_cell1
else:
        print('8th Testcase 1st checkpoint failed')
        sh.cell(row=9, column=3).value = 'Failed'
        fill_cell1 = PatternFill(patternType='solid',
                                 fgColor='FC2C03')
        sh.cell(row=9, column=3).fill = fill_cell1
time.sleep(5)
# ***********************************************************************************
print("Executing 9th TestCase")
sh.cell(row=10, column=1).value = 'TC_009_SearchByNpiNumber_Negative'
sh.cell(row=10, column=2).value = 'Verifying Search by NPI number functionality using invalid data'

# Click on toggle button(NPI Search level)
driver.find_element(By.XPATH, "//button[@role='switch']").click()
time.sleep(2)
# Enter invalid NPI Number
driver.find_element(By.XPATH, "//input[@placeholder='Enter NPI Number']").send_keys("18614777")
# Click on search
driver.find_element(By.XPATH, "//*[@id='root']/div[2]/div[2]/div/div[2]/button").click()
time.sleep(2)
# Capture error message
NPI_errormsg = driver.find_element(By.XPATH, "//*[@id='root']/div[2]/div[1]/div[3]/div/div/p")
print(NPI_errormsg.text)
Grid_errormsg = driver.find_element(By.XPATH, "//*[@id='root']/div[2]/div[3]/div[2]/span")
print(Grid_errormsg.text)
time.sleep(5)
# click on clear button
driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[2]/div[1]/div[1]/button[1]").click()
time.sleep(2)
# Click on search button
driver.find_element(By.XPATH, "//body[1]/div[1]/div[2]/div[2]/div[1]/div[2]/button[1]").click()
time.sleep(2)
# Capture error message
NPI_errormsg2 = driver.find_element(By.XPATH, "//*[@id='root']/div[2]/div[1]/div[3]/div/div/p")
print(NPI_errormsg2.text)
Grid_errormsg2 = driver.find_element(By.XPATH, "//*[@id='root']/div[2]/div[3]/div[2]/span")
print(Grid_errormsg2.text)
time.sleep(2)

if (NPI_errormsg.text) and (NPI_errormsg2.text) == 'Please enter valid NPI Number!' and (Grid_errormsg.text) and (Grid_errormsg2.text)=='Please Search Something...':
    print('TestCase: Search by NPI number with invalid data passed')
    print("8th Testcase passed")
    sh.cell(row=10, column=3).value = 'Passed'
    fill_cell2 = PatternFill(patternType='solid',
                             fgColor='35FC03')
    sh.cell(row=10, column=3).fill = fill_cell2
else:
    print('TestCase: Search by NPI number with invalid data failed')
    sh.cell(row=10, column=3).value = 'Failed'
    fill_cell1 = PatternFill(patternType='solid',
                             fgColor='FC2C03')
    sh.cell(row=10, column=3).fill = fill_cell1
time.sleep(3)
print("Saving workbook")
# saving a workbook
wrkbk.save("C:\\Users\\Sindhura\\Desktop\\Work\\Testing\\Automation\\Data_driven.xlsx" )
driver.close()
